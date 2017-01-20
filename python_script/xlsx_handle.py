import sys
from openpyxl import load_workbook

def handle_user_input(filename, uname_dir):

    wb = load_workbook(filename)
    sheets = wb.get_sheet_names()
    sheet0 = sheets[0]
    ws = wb.get_sheet_by_name(sheet0)
    #cols = ws.columns
    
    system_name = None
    sw_image_row = None
    sw_image_col = None

    for i in range(2, 10):
        for j in range(9, 30):
            data = str(ws.cell(row = j, column = i).value).strip()
            if data == "System Name":
               system_name = str(ws.cell(row = j+1, column = i).value).strip()
            if data == "sw-image":
               current_sw_image_name = str(ws.cell(row = j, column = i+2).value).strip()
               sw_image_row = j
               sw_image_col = i+2
               break
        else:
            continue
        break

    #if system_name == None or sw_image_row == None or sw_image_col == None:
    #   return
    print current_sw_image_name
    print system_name, sw_image_row, sw_image_col 

    ws.cell(row = sw_image_row, column = sw_image_col).value = current_sw_image_name + "_auto_" + uname_dir
    wb.save(filename)


# main function
# argument1: user.xlsx
# argument2: Xiao.A.Xu_alcatel-sbell.com.cn

print sys.argv[1]
handle_user_input(sys.argv[1], sys.argv[2])
