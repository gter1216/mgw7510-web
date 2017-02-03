from __future__ import division
from mgw7510.models import WebUser
from openpyxl import load_workbook
from pexpect import pxssh
import ce_deploy_scripts
import commands
import pexpect
import logging
import shutil
import time
import re
import os


def get_pak_list(pak_ip,
                 pak_username,
                 pak_passwd,
                 pak_path):

    try:
        s = pxssh.pxssh()
        s.login(pak_ip, pak_username, pak_passwd, original_prompt='[$#>]')
        cmd1 = 'cd ' + pak_path
        s.sendline(cmd1)
        s.prompt()
        s.sendline('find ./ -name *qcow2')
        s.prompt()
        result = s.before
        s.logout()

        final_result = {'ok': re.findall(r'nokia.*?qcow2', result)}
        return final_result
    except pxssh.ExceptionPxssh, e:
        return {'nok': 'Failed login to pak server, please check settings!'}


def update_progress_bar(user, data):
    user.progressBarData = data
    user.save()


def clean_work(user):
    if user.swImageName:
        uname_dir = user.tmpPath
        seedvm_info = {'ip': user.seedVMIp,
                       'username': user.seedVMUsername,
                       'passwd': user.seedVMPasswd,
                       'prompt': '#',
                       'openrc': user.seedVMOpenrcAbsPath,
                       'keypath': user.seedVMKeypairAbsPath,
                       'userdir': ce_deploy_scripts.SEEDVM_WORK_DIR + "/" + uname_dir}
        delete_image(seedvm_info, user.swImageName)


def deployment_success(user, perform_clean_work):
    if perform_clean_work == "yes":
        # do clean work
        clean_work(user)

    update_progress_bar(user, "100")

    logging.info('\nDeployment Successful! Remember to install scripts on V7510! \n')

    # shutdown logger
    logging.shutdown()


def deployment_failed(user, perform_clean_work):
    if perform_clean_work == "yes":
        # do clean work
        clean_work(user)

    update_progress_bar(user, "101")

    logging.info('\nDeployment failed, please check ERROR information \n')

    # shutdown logger
    logging.shutdown()


# general function to ssh to the specific host
def create_ssh_session(host, username, password, prompt):
    cmd = 'ssh ' + username + '@' + host

    logging.info('\n%s \n' % cmd)

    child = pexpect.spawn(cmd)
    ret = child.expect([pexpect.TIMEOUT, 'password:', 'Are you sure you want to continue connecting'], timeout=20)
    if ret == 0:
        raise Exception("\nssh to %s timeout \n" % host)
    elif ret == 1:
        child.sendline(password)
        child.expect(prompt)
    elif ret == 2:
        child.sendline("yes")
        ret = child.expect([pexpect.TIMEOUT, 'password'], timeout=20)
        if ret == 0:
            raise Exception("\nssh to %s timeout \n" % host)
        elif ret == 1:
            child.sendline(password)
            child.expect(prompt)

    # successful login to the host
    logging.info('%s \n' % child.before)
    return child


def update_progbar_by_scp(user, session, prompt, timer, progbar_total_incr):

    inc = float(progbar_total_incr) / 100

    session.expect("1%")
    logging.info('\n%s \n' % session.before)

    final_result_prev = 0
    initial_progress = int(user.progressBarData)

    while True:
        time.sleep(timer)
        session.expect("%")
        result1 = session.before
        result2 = result1.split()
        result3 = result2[5]
        final_result_second = int((int(result3)) * inc)
        if final_result_second > final_result_prev:
            logging.info('\n%s%%\n' % result1)
            user.progressBarData = str(initial_progress + final_result_second)
            user.save()
            final_result_prev = final_result_second
        if result3 == "100":
            session.expect(prompt)
            break

    logging.info('\n%s \n' % session.before)


def get_system_name(ws, j_row, i_col):
    for j in range(j_row, j_row+10):
        for i in range(i_col, i_col+10):
            data = str(ws.cell(row=j, column=i).value).strip()
            if data == "System Name":
                return str(ws.cell(row=j + 1, column=i).value).strip()


def get_sw_image(ws, j_row, i_col):
    for j in range(j_row, j_row+10):
        for i in range(i_col, i_col+10):
            data = str(ws.cell(row=j, column=i).value).strip()
            if data == "sw-image":
                return str(ws.cell(row=j, column=i + 2).value).strip(), j, i+2


def get_scm_ex_ip(ws, j_row, i_col):
    for j in range(j_row, j_row+10):
        for i in range(i_col, i_col+10):
            data = str(ws.cell(row=j, column=i).value).strip()
            if data == "IPv4 Fixed IP1":
                return str(ws.cell(row=j+1, column=i).value).strip(), str(ws.cell(row=j+1, column=i+1).value).strip()


def get_scm_ip(ws, j_row, i_col):
    for j in range(j_row, j_row+10):
        for i in range(i_col, i_col+10):
            data = str(ws.cell(row=j, column=i).value).strip()
            if data == "SCM OAM":
                return str(ws.cell(row=j, column=i + 2).value).strip()


def handle_user_input(filename, uname_dir):
    try:
        sheet_name = None
        system_name = None
        sw_image_row = None
        sw_image_col = None
        current_sw_image_name = None
        new_sw_image_name = None
        scm_ex_ip1 = None
        scm_ex_ip2 = None
        scm_oam_ip = None

        wb = load_workbook(filename)
        sheets = wb.get_sheet_names()
        sheet_name = sheets[0]
        if sheet_name is None:
            raise Exception("\nsheet name is None \n")
        # work sheet
        ws = wb.get_sheet_by_name(sheet_name)

        for i in range(2, 10):
            for j in range(9, 300):
                data = str(ws.cell(row=j, column=i).value).strip()
                if data == "System Parameters":
                    system_name = get_system_name(ws, j, i)
                elif data == "Deploy General Parameters":
                    (current_sw_image_name, sw_image_row, sw_image_col) = get_sw_image(ws, j, i)
                elif data == "Deploy External IP Parameters":
                    (scm_ex_ip1, scm_ex_ip2) = get_scm_ex_ip(ws, j, i)
                elif data == "IPv4 Interface":
                    scm_oam_ip = get_scm_ip(ws, j, i)
                    break
            else:
                continue
            break

        logging.info('\nsheet name is %s \n' % sheet_name)
        logging.info('\nsystem name is %s \n' % system_name)
        logging.info('\ncurrent sw image name is %s \n' % current_sw_image_name)
        logging.info('\ncurrent scm ext ip1 is %s \n' % scm_ex_ip1)
        logging.info('\ncurrent scm ext ip2 is %s \n' % scm_ex_ip2)
        logging.info('\ncurrent scm oam ip is %s \n' % scm_oam_ip)

        if system_name is None:
            raise Exception("\nsystem name is None \n")
        elif sw_image_row is None or sw_image_col is None:
            raise Exception("\ncurrent sw image is None \n")
        elif scm_ex_ip1 is None or scm_ex_ip2 is None or scm_oam_ip is None:
            raise Exception("\nscm ip is None\n")

        new_sw_image_name = current_sw_image_name + "_auto_" + uname_dir
        ws.cell(row=sw_image_row, column=sw_image_col).value = new_sw_image_name
        wb.save(filename)

        logging.info('\nnew sw image name is %s \n' % new_sw_image_name)
        return sheet_name, system_name, new_sw_image_name, scm_ex_ip1, scm_ex_ip2, scm_oam_ip

    except Exception, e:
        logging.error('\nproblem during handle user input file: %s \n' % str(e))
        return

# def update_progbar_by_scp(user, session, prompt, scp_step, progbar_total_incr):
#
#     scp_list = range(1, 100, scp_step)
#     scp_len = len(scp_list)
#
#     progbar_step = round(progbar_total_incr/(scp_len+1))
#     progbar_step = int(progbar_step)
#
#     scp_to = round(3600/(scp_len+1))
#     scp_to = int(scp_to)
#
#     # [1, 4, ... , 97]
#     for i_scp in scp_list:
#         scp_exp = str(i_scp)+"%"
#         session.expect(scp_exp, timeout=scp_to)
#         logging.info('\n%s \n' % session.before)
#         user.progressBarData = str(int(user.progressBarData) + progbar_step)
#         user.save()
#
#     session.expect(prompt, timeout=scp_to)
#     logging.info('\n%s \n' % session.before)
#     user.progressBarData = str(int(user.progressBarData) + progbar_step)
#     user.save()


def get_qcow2_md5_from_pak(pak_server_info, qcow2_name, select_rel):
    try:
        pak_server_ip = pak_server_info["ip"]
        pak_server_username = pak_server_info["username"]
        pak_server_password = pak_server_info["passwd"]
        pak_server_fp = pak_server_info["fp"]
        pak_prompt = pak_server_info["prompt"]

        pak_session = create_ssh_session(pak_server_ip, pak_server_username, pak_server_password, pak_prompt)

        pak_cmd = "cd " + "`" + "find " + pak_server_fp + "/" + "7510" + select_rel.replace(".", "") + "/ " + \
                  "-name " + qcow2_name + " -exec " + "dirname {} " + '\\;' + "`"

        pak_session.sendline(pak_cmd)
        pak_session.expect(pak_prompt)
        logging.info('\n%s \n' % pak_session.before)

        # do not need caculate MD5, comment the code
        # pak_session.sendline("md5sum " + qcow2_name)
        # pak_session.expect(pak_prompt)
        # result = pak_session.before
        # logging.info('\n%s \n' % result)
        # result = result.split()
        # pak_session.close()
        # return result[2]

        pak_session.sendline("more " + qcow2_name + ".md5")
        pak_session.expect(pak_prompt)
        result = pak_session.before
        logging.info('\n%s \n' % result)
        result = result.split()
        pak_session.close()
        return result[2]

    except Exception, e:
        logging.error('problem during ssh to pak server: %s \n' % str(e))
        return


def get_seedvm_qcow2_cached_flag_and_create_image(uname_dir, seedvm_info, qcow2_name, qcow2_md5, sw_image_name):
    try:
        seedvm_ip = seedvm_info["ip"]
        seedvm_username = seedvm_info["username"]
        seedvm_passwd = seedvm_info["passwd"]
        seedvm_prompt = seedvm_info["prompt"]

        seedvm_session = create_ssh_session(seedvm_ip, seedvm_username, seedvm_passwd, seedvm_prompt)

        # upload script seedvm_cache_check.sh to seedvm /root/
        sh_file_path = ce_deploy_scripts.BASE_DIR + "/shell_script/seedvm_cache_check.sh"
        seedvm_scp_cmd = "scp -r " + ce_deploy_scripts.WEB_SERVER_USERNAME + "@" + \
                         ce_deploy_scripts.WEB_SERVER_IP + ":" + sh_file_path + " /root/"
        seedvm_session.sendline(seedvm_scp_cmd)
        seedvm_ret = seedvm_session.expect([pexpect.TIMEOUT, '[p|P]assword:', 'connecting (yes/no)?'], timeout=100)
        if seedvm_ret == 0:
            raise Exception("\nscp to webserver timeout \n")
        elif seedvm_ret == 1:
            seedvm_session.sendline(ce_deploy_scripts.WEB_SERVER_PASSWORD)
        elif seedvm_ret == 2:
            logging.info('\n%s \n' % seedvm_session.before)
            seedvm_session.sendline("yes")
            seedvm_ret = seedvm_session.expect([pexpect.TIMEOUT, '[p|P]assword'], timeout=100)
            if seedvm_ret == 0:
                raise Exception("\nscp to webserver timeout \n")
            elif seedvm_ret == 1:
                logging.info('\n%s \n' % seedvm_session.before)
                seedvm_session.sendline(ce_deploy_scripts.WEB_SERVER_PASSWORD)
        seedvm_session.expect(seedvm_prompt)
        logging.info('\n%s \n' % seedvm_session.before)

        # perform shell script
        # $1 ===> /root/auto_ce_deploy
        # $2 ===> qcow2_name: nokia_a.qcow2
        # $3 ===> qcow2_md5: d41d8cd98f00b204e9800998ecf8427e
        # $4 ===> disk_limit: 15
        # $5 ===> source_file: /root/cloud-env/Rainbow-openrc.sh
        seedvm_cmd = "sh /root/seedvm_cache_check.sh " + "/root/auto_ce_deploy " + qcow2_name + " " + \
                     qcow2_md5 + " " + ce_deploy_scripts.SEEDVM_DISK_LIMIT + " " + seedvm_info["openrc"] + \
                     " " + uname_dir + " " + sw_image_name
        logging.info('\n%s \n' % seedvm_cmd)
        seedvm_session.sendline(seedvm_cmd)
        seedvm_session.expect(seedvm_prompt, timeout=200)
        logging.info('\n%s \n' % seedvm_session.before)
        seedvm_cmd = "echo $?"
        seedvm_session.sendline(seedvm_cmd)
        seedvm_session.expect(seedvm_prompt)
        result = seedvm_session.before
        logging.info('\n%s \n' % result)
        exitcode = result.split("\r\n")
        exitcode = exitcode[1]
        logging.info('\n exit code is: %s \n' % exitcode)

        # remove shell script
        seedvm_cmd = "rm -rf /root/seedvm_cache_check.sh"
        seedvm_session.sendline(seedvm_cmd)
        seedvm_session.expect(seedvm_prompt)
        logging.info('\n%s \n' % seedvm_session.before)

        seedvm_session.close()

        exitcode = int(exitcode)
        if exitcode == 1:
            logging.info('\nno cached qcow2 found on seedvm \n')
            return False
        elif exitcode == 2:
            logging.error('\nno enough disk storage on seedvm, please check! \n')
            return None
        elif exitcode == 3:
            logging.error('\ncreate image failed, please check! \n')
            return None
        elif exitcode == 4:
            logging.info('\nfind cached qcow2 on seedvm \n')
            return True
        elif exitcode == 5:
            logging.error('\ncreate image timeout, please check! \n')
            return None
        else:
            logging.error('\nunkown error \n')
            return None

    except Exception, e:
        logging.error('\nproblem during ssh to seedvm server: %s \n' % str(e))
        return None


def get_webserver_qcow2_cached_flag(qcow2_name, qcow2_md5):
    shell_file_path = ce_deploy_scripts.BASE_DIR + "/shell_script/webserver_cache_check.sh"

    ce_deploy_dir = ce_deploy_scripts.BASE_DIR + "/cache_dir/ce_deploy"

    shell_cmd = "sh " + shell_file_path + " " + ce_deploy_dir + " " + \
                qcow2_name + " " + qcow2_md5 + " " + ce_deploy_scripts.WEB_SERVER_DISK_LIMIT

    logging.info('\n%s\n' % shell_cmd)

    output = commands.getstatusoutput(shell_cmd)
    exitcode = (output[0]) >> 8
    logging.info('\n%s\n' % exitcode)
    logging.info('\n%s\n' % output[1])

    if exitcode == 1:
        logging.info('\nno cached qcow2 found on webserver \n')
        return False
    elif exitcode == 2:
        logging.error('\nno enough disk storage on webserver, please check! \n')
        return None
    elif exitcode == 3:
        logging.info('\nfind cached qcow2 on seedvm \n')
        return True
    else:
        logging.error('\nunkown error \n')
        return None


def download_files_to_webserver(user, pak_server_info, select_rel, select_pak, user_upload_dir, both_flag):
    try:
        pak_server_ip = pak_server_info["ip"]
        pak_server_username = pak_server_info["username"]
        pak_server_password = pak_server_info["passwd"]
        pak_server_fp = pak_server_info["fp"]
        pak_prompt = pak_server_info["prompt"]

        pak_session = create_ssh_session(pak_server_ip, pak_server_username, pak_server_password, pak_prompt)

        # cd `find /viewstores/public/SLP/7510C71/ -name
        # nokia-mgw-rhel7.2-3.10.0-327.18.2.ae1a3116.x86_64.qcow2 -exec dirname {} \;`
        pak_cmd = "cd " + "`" + "find " + pak_server_fp + "/" + "7510" + select_rel.replace(".", "") + "/ " + \
                  "-name " + select_pak + " -exec " + "dirname {} " + '\\;' + "`"

        pak_session.sendline(pak_cmd)
        pak_session.expect(pak_prompt)
        logging.info('\n%s \n' % pak_session.before)

        logging.info('\nls \n')
        pak_session.sendline('ls')
        pak_session.expect(pak_prompt)
        logging.info('\n%s \n' % pak_session.before)

        if both_flag is True:
            # scp -r *M_O*csar.zip *.qcow2 root@135.251.216.181:user_upload_dir
            pak_cmd_qcow2 = "scp -r " + "*.qcow2 " + ce_deploy_scripts.WEB_SERVER_USERNAME + "@" + \
                            ce_deploy_scripts.WEB_SERVER_IP + ":" + ce_deploy_scripts.BUFFER_DIR
            logging.info('\n%s \n' % pak_cmd_qcow2)
            pak_session.sendline(pak_cmd_qcow2)
            pak_ret = pak_session.expect([pexpect.TIMEOUT, '[p|P]assword:', 'connecting (yes/no)?'], timeout=5)
            if pak_ret == 0:
                raise Exception("\nscp to webserver timeout \n")
            elif pak_ret == 1:
                pak_session.sendline(ce_deploy_scripts.WEB_SERVER_PASSWORD)
            elif pak_ret == 2:
                logging.info('\n%s \n' % pak_session.before)
                pak_session.sendline("yes")
                pak_ret = pak_session.expect([pexpect.TIMEOUT, '[p|P]assword'], timeout=5)
                if pak_ret == 0:
                    raise Exception("\nscp to webserver timeout \n")
                elif pak_ret == 1:
                    logging.info('\n%s \n' % pak_session.before)
                    pak_session.sendline(ce_deploy_scripts.WEB_SERVER_PASSWORD)

            update_progbar_by_scp(user=user, session=pak_session, prompt=pak_prompt,
                                  timer=0.1, progbar_total_incr=35)

            # download csar and qcow2 successfull
            # move qcow2 from buffer dir to cache dir
            cmd = "mv -f " + ce_deploy_scripts.BUFFER_DIR + "/" + select_pak + " " + \
                  ce_deploy_scripts.CACHE_DIR
            logging.info('\n%s \n' % cmd)
            os.system(cmd)

        pak_cmd_csar = "scp -r " + "*M_O*csar.zip yact-C* yact-nokia-mgw* " + ce_deploy_scripts.WEB_SERVER_USERNAME + "@" + \
                       ce_deploy_scripts.WEB_SERVER_IP + ":" + user_upload_dir

        logging.info('\n%s \n' % pak_cmd_csar)
        pak_session.sendline(pak_cmd_csar)
        pak_ret = pak_session.expect([pexpect.TIMEOUT, '[p|P]assword:', 'connecting (yes/no)?'], timeout=5)
        if pak_ret == 0:
            raise Exception("\nscp to webserver timeout \n")
        elif pak_ret == 1:
            pak_session.sendline(ce_deploy_scripts.WEB_SERVER_PASSWORD)
        elif pak_ret == 2:
            logging.info('\n%s \n' % pak_session.before)
            pak_session.sendline("yes")
            pak_ret = pak_session.expect([pexpect.TIMEOUT, '[p|P]assword'], timeout=5)
            if pak_ret == 0:
                raise Exception("\nscp to webserver timeout \n")
            elif pak_ret == 1:
                logging.info('\n%s \n' % pak_session.before)
                pak_session.sendline(ce_deploy_scripts.WEB_SERVER_PASSWORD)
        pak_session.expect(pak_prompt)
        logging.info('\n%s \n' % pak_session.before)
        pak_session.close()
        return True

    except Exception, e:
        logging.error('\nproblem during ssh to pak server: %s \n' % str(e))
        return False


def delete_image(seedvm_info, sw_image_name):
    try:
        seedvm_ip = seedvm_info["ip"]
        seedvm_username = seedvm_info["username"]
        seedvm_passwd = seedvm_info["passwd"]
        seedvm_prompt = seedvm_info["prompt"]
        seedvm_openrc = seedvm_info["openrc"]

        seedvm_session = create_ssh_session(seedvm_ip, seedvm_username, seedvm_passwd, seedvm_prompt)
        logging.info('\ndelete image \n')

        seedvm_session.sendline("source " + seedvm_openrc)
        seedvm_session.expect(seedvm_prompt)
        logging.info('\n%s \n' % seedvm_session.before)

        delete_glance_cmd = "glance image-delete " + sw_image_name
        logging.info('\n%s \n' % delete_glance_cmd)
        seedvm_session.sendline(delete_glance_cmd)
        seedvm_session.expect(seedvm_prompt, timeout=50)

        logging.info('\ndelete glance image successful \n')
        seedvm_session.close()
        return True

    except Exception, e:
        logging.error('\nproblem during ssh to seedvm server: %s \n' % str(e))
        return False


def upload_qcow2_to_seed_create_image(user, seedvm_info, select_pak, sw_image_name):
    try:
        seedvm_ip = seedvm_info["ip"]
        seedvm_username = seedvm_info["username"]
        seedvm_passwd = seedvm_info["passwd"]
        seedvm_prompt = seedvm_info["prompt"]
        seedvm_openrc = seedvm_info["openrc"]

        seedvm_session = create_ssh_session(seedvm_ip, seedvm_username, seedvm_passwd, seedvm_prompt)

        # upload qcow2 to buffer dir

        seedvm_cmd_scp = "scp -r " + ce_deploy_scripts.WEB_SERVER_USERNAME + "@" + \
                         ce_deploy_scripts.WEB_SERVER_IP + ":" + ce_deploy_scripts.CACHE_DIR + "/" + \
                         select_pak + " " + ce_deploy_scripts.SEEDVM_BUFFER_dIR

        logging.info('\n%s \n' % seedvm_cmd_scp)

        seedvm_session.sendline(seedvm_cmd_scp)

        seedvm_ret = seedvm_session.expect([pexpect.TIMEOUT, '[p|P]assword:', 'connecting (yes/no)?'], timeout=100)
        if seedvm_ret == 0:
            raise Exception("\nscp to webserver timeout \n")
        elif seedvm_ret == 1:
            seedvm_session.sendline(ce_deploy_scripts.WEB_SERVER_PASSWORD)
        elif seedvm_ret == 2:
            logging.info('\n%s \n' % seedvm_session.before)
            seedvm_session.sendline("yes")
            seedvm_ret = seedvm_session.expect([pexpect.TIMEOUT, '[p|P]assword'], timeout=100)
            if seedvm_ret == 0:
                raise Exception("\nscp to webserver timeout \n")
            elif seedvm_ret == 1:
                logging.info('\n%s \n' % seedvm_session.before)
                seedvm_session.sendline(ce_deploy_scripts.WEB_SERVER_PASSWORD)

        update_progbar_by_scp(user=user,
                              session=seedvm_session,
                              prompt=seedvm_prompt,
                              timer=0.1,
                              progbar_total_incr=35)

        logging.info('\nmove qcow2 from buffer to cache, on seedvm \n')
        move_cmd = "mv -f " + ce_deploy_scripts.SEEDVM_BUFFER_dIR + "/" + select_pak + " " + ce_deploy_scripts.SEEDVM_CACHE_dIR + "/" + select_pak
        logging.info('\n%s \n' % move_cmd)
        seedvm_session.sendline(move_cmd)
        seedvm_session.expect(seedvm_prompt)
        logging.info('\n%s \n' % seedvm_session.before)

        # ================= Create Image
        logging.info('\ncreate image \n')
        seedvm_session.sendline("cd " + ce_deploy_scripts.SEEDVM_BUFFER_dIR)
        seedvm_session.expect(seedvm_prompt)
        logging.info('\n%s \n' % seedvm_session.before)

        seedvm_session.sendline("source " + seedvm_openrc)
        seedvm_session.expect(seedvm_prompt)
        logging.info('\n%s \n' % seedvm_session.before)

        #glance_name = select_pak.strip(".qcow2") + "_auto_" + uname_dir

        create_glance_cmd = "glance image-create --name=" + sw_image_name + " --file=" + select_pak + \
                            " --disk-format=qcow2  --container-format=bare  --is-public=false --is-protected=false"
        logging.info('\n%s \n' % create_glance_cmd)
        seedvm_session.sendline(create_glance_cmd)
        seedvm_ret = seedvm_session.expect(['Errno', seedvm_prompt], timeout=50)
        if seedvm_ret == 0:
            logging.info('\n%s \n' % seedvm_session.before)
            raise Exception("\ncreate glance image failed \n")
        elif seedvm_ret == 1:
            logging.info('\n%s \n' % seedvm_session.before)

        logging.info('\ncreate glance image successful \n')
        seedvm_session.close()
        return True

    except Exception, e:
        logging.error('\nproblem during ssh to seedvm server: %s \n' % str(e))
        return False


def make_yaml_scripts(uname_dir, sheet_name):
    try:
        shell_file_path = ce_deploy_scripts.BASE_DIR + "/shell_script/make_yaml_script.sh"

        shell_cmd = shell_file_path + " " + uname_dir + " " + sheet_name

        logging.info('\n%s \n' % shell_cmd)

        output = commands.getstatusoutput(shell_cmd)

        logging.info('\n%s\n' % output[1])
        if output[0] != 0:
            raise Exception("\nmake yaml and script failed \n")
        return True
    except Exception, e:
        logging.error('\nproblem during make yaml & script: %s \n' % str(e))
        return False

# def check_stack_final_result(seedvm_session, seedvm_prompt, system_name):
#     while()


def check_stack_delete_status(seedvm_session, seedvm_prompt, system_name, timeout):
    t0 = time.time()

    seedvm_list_cmd = "heat stack-list "

    while True:
        seedvm_session.sendline(seedvm_list_cmd)
        seedvm_session.expect(seedvm_prompt)
        list_result = seedvm_session.before

        logging.info('\n%s \n' % list_result)

        if re.findall(system_name, list_result) == []:
            return True

        if time.time() - t0 > timeout:
            break

        time.sleep(5)

    return False


def check_stack_final_result(seedvm_session, seedvm_prompt, system_name, timeout):
    t0 = time.time()
    cmd = "heat stack-list"

    while True:
        seedvm_session.sendline(cmd)
        seedvm_session.expect(seedvm_prompt)
        result = seedvm_session.before
        logging.info('\n%s \n' % result)
        result = result.split()
        index = result.index(system_name)
        create_state = result[index+2]
        if create_state == "CREATE_COMPLETE":
            return True
        elif create_state == "CREATE_FAILED":
            break
        if time.time() - t0 > timeout:
            break
        time.sleep(10)

    return False


def create_stack(
        seedvm_info,
        user_upload_dir,
        system_name,
        scm_ex_ip1,
        scm_ex_ip2,
        scm_oam_ip):

    try:
        seedvm_ip = seedvm_info["ip"]
        seedvm_username = seedvm_info["username"]
        seedvm_passwd = seedvm_info["passwd"]
        seedvm_prompt = seedvm_info["prompt"]
        seedvm_openrc = seedvm_info["openrc"]
        seedvm_keypath = seedvm_info["keypath"]

        seedvm_session = create_ssh_session(seedvm_ip, seedvm_username, seedvm_passwd, seedvm_prompt)

        seedvm_user_dir = seedvm_info["userdir"]

        # create user dir on seedvm if not exist
        # if exist, clean the content
        seedvm_session.sendline("mkdir " + seedvm_user_dir)
        seedvm_session.expect(seedvm_prompt)
        logging.info('\n%s \n' % seedvm_session.before)

        seedvm_session.sendline("cd " + seedvm_user_dir + " && " + "rm -rf " + " * ")
        seedvm_session.expect(seedvm_prompt)
        logging.info('\n%s \n' % seedvm_session.before)

        csar_dir = user_upload_dir + "/*.csar/"

        seedvm_cmd = "scp -r " + ce_deploy_scripts.WEB_SERVER_USERNAME + "@" + ce_deploy_scripts.WEB_SERVER_IP + \
                     ":" + csar_dir + " ./ "

        logging.info('\n%s \n' % seedvm_cmd)

        seedvm_session.sendline(seedvm_cmd)
        seedvm_ret = seedvm_session.expect([pexpect.TIMEOUT, '[p|P]assword:', 'connecting (yes/no)?'], timeout=300)

        if seedvm_ret == 0:
            raise Exception("\nscp to webserver timeout \n")
        elif seedvm_ret == 1:
            seedvm_session.sendline(ce_deploy_scripts.WEB_SERVER_PASSWORD)
        elif seedvm_ret == 2:
            logging.info('\n%s \n' % seedvm_session.before)
            seedvm_session.sendline("yes")
            seedvm_ret = seedvm_session.expect([pexpect.TIMEOUT, '[p|P]assword'], timeout=300)
            if seedvm_ret == 0:
                raise Exception("\nscp to webserver timeout \n")
            elif seedvm_ret == 1:
                logging.info('\n%s \n' % seedvm_session.before)
                seedvm_session.sendline(ce_deploy_scripts.WEB_SERVER_PASSWORD)

        seedvm_session.expect(seedvm_prompt, timeout=300)
        logging.info('\n%s \n' % seedvm_session.before)

        logging.info('\nStep7: generate hot and env files\n')

        seedvm_session.sendline("cd *.csar/scripts")
        seedvm_session.expect(seedvm_prompt)
        logging.info('\n%s \n' % seedvm_session.before)

        # ./template.py -a deploy -r cloud_config/cloud-resource-data.yaml
        seedvm_session.sendline("./template.py -a deploy -r cloud_config/cloud-resource-data.yaml")
        seedvm_session.expect(seedvm_prompt)

        # get heat create command
        logging.info('\nStep8: create stack\n')
        cmd_result = seedvm_session.before
        logging.info('\n%s \n' % cmd_result)
        cmd_result = re.findall(r'heat.*\r', cmd_result)
        create_heat_cmd = cmd_result[0].strip('\r')
        logging.info('\n%s \n' % create_heat_cmd)

        # create 7510 stack
        seedvm_session.sendline("source " + seedvm_openrc)
        seedvm_session.expect(seedvm_prompt)
        logging.info('\n%s \n' % seedvm_session.before)

        seedvm_list_cmd = "heat stack-list "
        seedvm_session.sendline(seedvm_list_cmd)
        seedvm_session.expect(seedvm_prompt)
        list_result = seedvm_session.before

        if re.findall(system_name, list_result) != []:
            logging.info('\nthere is an existed same name stack, delete it\n')

            # delete old same name stack
            seedvm_del_cmd = "heat stack-delete " + system_name
            seedvm_session.sendline(seedvm_del_cmd)
            seedvm_session.expect(seedvm_prompt)
            logging.info('\n%s \n' % seedvm_session.before)

            if check_stack_delete_status(
                    seedvm_session, seedvm_prompt, system_name, timeout=120) is not True:
                raise Exception("\nthere is still an existed same name stack, failed \n")

        seedvm_session.sendline(create_heat_cmd)
        seedvm_session.expect(seedvm_prompt, timeout=300)
        logging.info('\n%s \n' % seedvm_session.before)

        seedvm_cmd = "echo $?"
        seedvm_session.sendline(seedvm_cmd)
        seedvm_session.expect(seedvm_prompt)
        stack_create_result = seedvm_session.before
        logging.info('\n%s \n' % stack_create_result)
        stack_create_result = stack_create_result.split("\r\n")
        stack_create_result_code = stack_create_result[1]
        logging.info('\n stack create result is: %s \n' % stack_create_result_code)

        if int(stack_create_result_code) != 0:
            logging.error('\ncreate stack failed \n')
            return False

        # check stack create final result by "heat stack-list | grep system_name"
        if check_stack_final_result(seedvm_session, seedvm_prompt, system_name, timeout=360) is not True:
            raise Exception("\nfinal create stack result is not CREATE_COMPLETE \n")


        # ======= step9: generate particular file "UUID.TXT"
        logging.info('\nStep9: generate particular file UUID.TXT \n')
        seedvm_cmd = "./stack.py -a get-uuid -s " + system_name + " -o ../bulk-config/UUID.TXT"
        seedvm_session.sendline(seedvm_cmd)
        seedvm_session.expect(seedvm_prompt, timeout=5)
        logging.info('\n%s \n' % seedvm_session.before)

        # ======= step10: put needed files to v7510
        logging.info('\nStep10: put needed files to v7510 \n')

        seedvm_session.sendline("cd ../bulk-config")
        seedvm_session.expect(seedvm_prompt)
        logging.info('\n%s \n' % seedvm_session.before)

        seedvm_session.sendline("rm -rf ~/.ssh/known_hosts")
        seedvm_session.expect(seedvm_prompt)
        logging.info('\n%s \n' % seedvm_session.before)

        cmd = "scp -i " + seedvm_keypath + " * " + "cloud-user@" + scm_ex_ip1 + ":/opt/v7510/data/"
        logging.info('\n%s \n' % cmd)
        seedvm_session.sendline(cmd)
        ret = seedvm_session.expect([pexpect.TIMEOUT,
                                     'Are you sure you want to continue connecting',
                                     seedvm_prompt],
                                    timeout=200)
        if ret == 0:
            logging.info('\n%s \n' % seedvm_session.before)
            raise Exception("\nscp to %s timeout \n" % scm_ex_ip1)
        elif ret == 1:
            seedvm_session.sendline("yes")
            seedvm_session.expect(seedvm_prompt, timeout=20)
            logging.info('\n%s \n' % seedvm_session.before)
        elif ret == 2:
            logging.info('\n%s \n' % seedvm_session.before)
        elif ret == 3:
            logging.info('\n%s \n' % seedvm_session.before)

        cmd = "scp -i " + seedvm_keypath + " * " + "cloud-user@" + scm_ex_ip2 + ":/opt/v7510/data/"
        logging.info('\n%s \n' % cmd)
        seedvm_session.sendline(cmd)
        ret = seedvm_session.expect([pexpect.TIMEOUT,
                                     'Are you sure you want to continue connecting',
                                     seedvm_prompt],
                                    timeout=200)
        if ret == 0:
            logging.info('\n%s \n' % seedvm_session.before)
            raise Exception("\nscp to %s timeout \n" % scm_ex_ip2)
        elif ret == 1:
            seedvm_session.sendline("yes")
            seedvm_session.expect(seedvm_prompt, timeout=20)
            logging.info('\n%s \n' % seedvm_session.before)
        elif ret == 2:
            logging.info('\n%s \n' % seedvm_session.before)

        # ==============================================================================
        # scripts running on V7510 TBD
        # ==============================================================================

        # # ======= step11: run instal script on v7510
        # logging.info('\nStep11: run instal script on v7510 \n')
        #
        # prompt = "\$"
        #
        # cmd = "ssh -i " + seedvm_keypath + " cloud-user@" + scm_ex_ip1
        # seedvm_session.sendline(cmd)
        # seedvm_session.expect(prompt, timeout=20)
        # logging.info('\n%s \n' % seedvm_session.before)
        #
        # seedvm_session.sendline("su root")
        # ret = seedvm_session.expect([pexpect.TIMEOUT, 'Password:'], timeout=20)
        # if ret == 0:
        #     raise Exception("\nssh to %s timeout \n" % scm_ex_ip1)
        # elif ret == 1:
        #     seedvm_session.sendline("-assured")
        #     seedvm_session.expect("#")
        #     logging.info('\n%s \n' % seedvm_session.before)
        #
        # seedvm_session.sendline("node-console -s 10")
        # seedvm_session.sendline("\n")
        # ret = seedvm_session.expect([pexpect.TIMEOUT, "Login:", "vMGx#"], timeout=20)
        # if ret == 0:
        #     raise Exception("\nnode-console -s 10 timeout \n")
        # elif ret == 1:
        #     seedvm_session.sendline("diag")
        #     seedvm_session.expect("Password:")
        #     seedvm_session.sendline("-assured")
        #     seedvm_session.expect("vMGx#")
        #     logging.info('\n%s \n' % seedvm_session.before)
        # elif ret == 2:
        #     logging.info('\n%s \n' % seedvm_session.before)
        #
        # seedvm_session.sendline("run script INSTALL0.SCR")
        # seedvm_session.expect("vMGx#", timeout=120)
        # result = seedvm_session.before
        # logging.info('\n%s \n' % result)
        # seedvm_session.close()
        #
        # # login to active scm board to run left script
        # logging.info('\nStep12: login to active scm board to run left script \n')
        # seedvm_session = create_ssh_session(seedvm_ip, seedvm_username, seedvm_passwd, seedvm_prompt)
        #
        # new_prompt = "vMGx#"
        # cmd = "ssh " + "diag@" + scm_oam_ip
        # seedvm_session.sendline(cmd)
        # ret = seedvm_session.expect([pexpect.TIMEOUT, 'password:', 'Are you sure you want to continue connecting'],
        #                             timeout=240)
        # if ret == 0:
        #     raise Exception("\nssh to %s timeout \n" % scm_oam_ip)
        # elif ret == 1:
        #     seedvm_session.sendline("-assured")
        #     seedvm_session.expect(new_prompt)
        # elif ret == 2:
        #     seedvm_session.sendline("yes")
        #     ret = seedvm_session.expect([pexpect.TIMEOUT, 'password'], timeout=20)
        #     if ret == 0:
        #         raise Exception("\nssh to %s timeout \n" % scm_oam_ip)
        #     elif ret == 1:
        #         seedvm_session.sendline("-assured")
        #         seedvm_session.expect(new_prompt)
        # seedvm_session.expect(new_prompt, timeout=20)
        # logging.info('\n%s \n' % seedvm_session.before)
        #
        # seedvm_session.sendline("run script INSTALL.SCR")
        # seedvm_session.expect(new_prompt, timeout=60)
        # logging.info('\n%s \n' % seedvm_session.before)
        #
        # seedvm_session.sendline("run script BULK.SCRR")
        # seedvm_session.expect(new_prompt, timeout=60)
        # logging.info('\n%s \n' % seedvm_session.before)

        seedvm_session.close()

    except Exception, e:
        logging.error('\nproblem during ssh to seedvm server: %s \n' % str(e))
        return False



